#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
NetserGroup - Acumulador de Datos Históricos
Lee Reporte_NetserGroup_Final.xlsx y agrega los datos del día al CSV histórico.
Evita duplicados por fecha.
Autor: Humberto Henriquez
"""

import csv
import os
import sys
from datetime import datetime

try:
    import openpyxl
except ImportError:
    print("[ERROR] Se requiere openpyxl. Instalar con: pip install openpyxl")
    sys.exit(1)

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
EXCEL_FILE = os.path.join(SCRIPT_DIR, "Reporte_NetserGroup_Final.xlsx")
CSV_CASOS = os.path.join(SCRIPT_DIR, "historico_casos.csv")
CSV_BOTS = os.path.join(SCRIPT_DIR, "historico_bots.csv")
SHEET_NAME = "Datos"

CLIENTS = [
    "HP Comercial", "HPE", "Payless", "Netapp", "Lexmark",
    "Lexmark Kit", "CTDI", "Monthly Fee", "Lenovo"
]

BOTS = [
    "BackUp Mobility", "Cierre POs", "Cierre Alpha", "Cierre HPCM",
    "Tasas Cambio", "Encuestas Dell", "Respaldo Invoice", "Cierre Residencias",
    "Receiving Lab", "Reporte Inv HP", "HPCM Cenam", "HPCM Chile",
    "Licencias FSM", "Regularizacion Mobility"
]


def safe_int(val):
    if val is None:
        return 0
    try:
        return int(float(val))
    except (ValueError, TypeError):
        return 0


def is_bot_ok(val):
    if val is None:
        return False
    s = str(val).strip()
    return s in ("\u2714", "\u2714\ufe0f", "OK", "ok", "1", "TRUE", "True", "true", "\u2713")


def parse_date(val):
    """Extrae fecha del valor de la celda."""
    if val is None:
        return datetime.now().strftime("%Y-%m-%d")
    if isinstance(val, datetime):
        return val.strftime("%Y-%m-%d")
    s = str(val).strip().strip("()")
    # Intentar parsear formatos comunes
    for fmt in ["%d/%m/%Y %H:%M:%S", "%d/%m/%Y", "%Y-%m-%d", "%m/%d/%Y"]:
        try:
            return datetime.strptime(s, fmt).strftime("%Y-%m-%d")
        except ValueError:
            continue
    return datetime.now().strftime("%Y-%m-%d")


def get_existing_dates(csv_path):
    """Lee fechas ya registradas en el CSV."""
    dates = set()
    if os.path.exists(csv_path):
        with open(csv_path, "r", encoding="utf-8") as f:
            reader = csv.DictReader(f)
            for row in reader:
                dates.add(row.get("Fecha", ""))
    return dates


def read_excel_data():
    """Lee datos del Excel y retorna diccionarios de casos y bots."""
    if not os.path.exists(EXCEL_FILE):
        print(f"[ERROR] No se encontro: {EXCEL_FILE}")
        sys.exit(1)

    wb = openpyxl.load_workbook(EXCEL_FILE, data_only=True)
    ws = wb[SHEET_NAME]

    # Mapear headers a columnas
    headers = {}
    for c in range(1, ws.max_column + 1):
        h = ws.cell(1, c).value
        if h:
            headers[h.strip()] = c

    # Leer fila 2 (datos del dia)
    fecha_val = ws.cell(2, headers.get("Fecha", 1)).value
    fecha = parse_date(fecha_val)

    # Casos por cliente
    casos = {"Fecha": fecha}
    total = 0
    for client in CLIENTS:
        col = headers.get(client)
        val = safe_int(ws.cell(2, col).value) if col else 0
        casos[client] = val
        total += val
    casos["Total"] = total

    # Estado de bots
    bots_data = {"Fecha": fecha}
    ok_count = 0
    for bot in BOTS:
        col = headers.get(bot)
        estado = is_bot_ok(ws.cell(2, col).value) if col else False
        bots_data[bot] = "OK" if estado else "FAIL"
        if estado:
            ok_count += 1
    bots_data["BotsOK"] = ok_count
    bots_data["BotsFail"] = len(BOTS) - ok_count
    bots_data["TotalBots"] = len(BOTS)

    wb.close()
    return casos, bots_data


def append_to_csv(csv_path, data, fieldnames):
    """Agrega una fila al CSV. Crea el archivo si no existe."""
    file_exists = os.path.exists(csv_path)

    with open(csv_path, "a", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        if not file_exists:
            writer.writeheader()
        writer.writerow(data)


def main():
    print("=" * 50)
    print("  NetserGroup - Actualizando Historico")
    print("=" * 50)

    casos, bots_data = read_excel_data()
    fecha = casos["Fecha"]
    print(f"\nFecha del reporte: {fecha}")

    # --- Historico de casos ---
    casos_fields = ["Fecha"] + CLIENTS + ["Total"]
    existing_casos = get_existing_dates(CSV_CASOS)

    if fecha in existing_casos:
        # Actualizar la fila existente
        rows = []
        with open(CSV_CASOS, "r", encoding="utf-8") as f:
            reader = csv.DictReader(f)
            for row in reader:
                if row["Fecha"] == fecha:
                    rows.append(casos)
                else:
                    rows.append(row)
        with open(CSV_CASOS, "w", newline="", encoding="utf-8") as f:
            writer = csv.DictWriter(f, fieldnames=casos_fields)
            writer.writeheader()
            writer.writerows(rows)
        print(f"[CASOS] Datos actualizados para {fecha}")
    else:
        append_to_csv(CSV_CASOS, casos, casos_fields)
        print(f"[CASOS] Nueva entrada agregada: {fecha}")

    # --- Historico de bots ---
    bots_fields = ["Fecha"] + BOTS + ["BotsOK", "BotsFail", "TotalBots"]
    existing_bots = get_existing_dates(CSV_BOTS)

    if fecha in existing_bots:
        rows = []
        with open(CSV_BOTS, "r", encoding="utf-8") as f:
            reader = csv.DictReader(f)
            for row in reader:
                if row["Fecha"] == fecha:
                    rows.append(bots_data)
                else:
                    rows.append(row)
        with open(CSV_BOTS, "w", newline="", encoding="utf-8") as f:
            writer = csv.DictWriter(f, fieldnames=bots_fields)
            writer.writeheader()
            writer.writerows(rows)
        print(f"[BOTS] Datos actualizados para {fecha}")
    else:
        append_to_csv(CSV_BOTS, bots_data, bots_fields)
        print(f"[BOTS] Nueva entrada agregada: {fecha}")

    # Resumen
    print(f"\n--- Resumen del dia ---")
    print(f"Total casos: {casos['Total']}")
    for c in CLIENTS:
        if casos[c] > 0:
            print(f"  {c}: {casos[c]}")
    print(f"Bots OK: {bots_data['BotsOK']}/{bots_data['TotalBots']}")
    print(f"\nArchivos:")
    print(f"  {CSV_CASOS}")
    print(f"  {CSV_BOTS}")
    print()


if __name__ == "__main__":
    main()
